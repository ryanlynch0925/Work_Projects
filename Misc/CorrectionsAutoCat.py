import pandas as pd
from openpyxl import load_workbook
import re

# Load the Excel file
data_path = r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Desktop\dev\Work_Projects\Expense_Reporting_Emails\Corrections V2.0.xlsx"
# data_path = r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\CatTest.xlsx"
# sheet_name = 'Cats'
df = pd.read_excel(data_path, sheet_name="Need New")
# df = pd.read_excel(data_path, sheet_name="Cats")

# Dictionary for categorizing notes
category_dict = {
    'Advertising & Promotion': ['dog treats', 'dog bones', 'candy for kids', 'fun pops', 'charity', 'baskets for fundraiser', 'lollipops', 'marketing', 'newspaper ad', 'parade', 'fundraiser', 'customer complaint', 'doggie treats', 'candy for customers', 'flowers for customers', 'suckers for customers', 'sign holder', 'brochures', 'ribon cutting', 'advertising', 'braves'],
    'Airfare': ['flight', 'airfare', 'bags', 'bags fee', 'bag check', 'no seats remained', 'baggage', 'check bag', 'round trip', 'air travel', 'travel insurance', 'luggage'],
    'Annual Inspection': ['boiler inspection', 'back flow', 'backflow'],
    'Auto Maintenance': ['air tire up', 'new tire', 'oil change', 'washer fluid'],
    'Auto Repairs': ['oil pan gaurd'],
    'Building Repairs': ['painter', 'hvac'],
    'Car Rental': ['car rental', 'rental truck', 'rental car', 'rental for'],
    'Car Wash': ['monthly wash', 'network test', 'testing xpt', 'competitor', 'competition', 'xpt transaction', 'xpt test', 'credit card test', 'drb test', 'drb remote', 'emv restart', 'other wash', 'car wash', 'test wash', 'drb testing', 'tidal wave', 'testing credit card', 'card reader test', 'test', 'sl training', 'it test', 'xpt testing', 'hub 2', 'fleet charge test'],
    'Charity Day': ['charity day'],
    'Cleaning Services': ['3rd party tunnel cleaning'],
    'Computer': [],
    'Dues & Subscriptions': ['sams club', 'subscription', 'prime membership'],
    'Drug Testing': ['drug'],
    'Electric Repairs': ['electrician'],
    'Employee Incentive - Entertainment': [],
    ################################### Employee Incentive - Meals ###################
    'Employee Incentive - Meals': ['donuts for team', 'food for', 'employee meal', 'pizza', 'lunch for', 'dinner for', 'team meals', 'coffee', 'cookies', 'cake', 'meal for', 'lunch', 'chips and dip', 'food to team', 'donuts', 'employee meals', 'chick fil a', 'breakfast', 'took sl out', 'dinner with', 'took sl and interim out', 'party meals'],
    ##################################################################################
    'Employee Incentives - Other': ['birthday present', 'contest', 'bday', 'google reviews', 'for contests', 'gift card for employee', 'employees thermas'],
    'Entertainment': [],
    'Equipment': ['large shop fan', 'tool chest', 'electric pressure washer', 'surface cleaner', 'washing machine', 'water heater', 'new washer'],
    'Equipment - Non-SL': [],
    'Equipment Repairs': ['3rd party rebuild'],
    'Fuel': ['fuel', 'gas', 'gasoline', 'unleaded', 'mileage'],
    ################################# Landscaping #################################
    'Landscaping': ['sprinklers', 'lawncare', 'flowers', 'landscaping', 'lawn service', 'weed killer', 'soil', 'lawn care', 'mowing', 'lawn mower', 'aeration', 'flower', 'planter', 'mulch', 'weed eater', 'fertilizer', 'potting mix', 'rocks', 'tree limbs', 'chainsaw', 'bi weekly', 'grubex', 'weedeater', 'weed pullers', 'butlers', 'yard dustpan', 'yard pan', 'lawn', 'weed', 'caladium', 'planters', 'transplanter', 'cultivator', 'trowel', 'plants', 'grass seed', 'landscaper', 'cobblestone', 'backpack blower', 'walkway stone', 'irrigation', 'landscape', 'round up', 'trimmer', 'Sprinkler system repair', 'bushes', 'trugreen', 'garden', 'landscaping', 'ground spikes', 'trufuel', 'grass cut', 'Top turf aerator', 'roundup', 'edging', 'biweekly cut', 'grounds keeping', 'groundskeeping'],
    ###############################################################################
    'Laundry & Linens': [],
    'Locksmith': ['locksmith',  'locksmiths'],
    'Lodging': ['lodging', 'hotel', 'rooming for', 'stay in', 'overnight', 'air bnb', 'trainer room', 'support trainer', 'room'],
    'Maintenance': [],
    'Marketing Services': [],
    'Miscellaneous Expense - Non-SL': ['personal', 'red bull', 'energy drink', 'charge back to my pay', 'dispute', 'accidentally purchased with card'],
    'Office Furniture': ['old refrigerator', 'shelving', 'cabinet', 'picnic bench', 'water dispenser', 'toaster oven', 'ice maker', 'storage rack', 'refrigerator', 'tv bracket', 'Office furniture', 'water cooler', 'icemaker', 'fridge', 'Storage shelves', 'desk top', 'shelf', 'water machine'],
    ################################# Office Supplies #################################
    'Office Supplies': ['bathroom', 'ink', 'key copy', 'keys', 'printer', 'copy paper', 'usb splitters', 'receipt', 'keyboard', 'mouse', 'toner', 'office', 'cooling   wraps', 'file folder', 'toilet paper', 'tp', 'cups', 'xpt paper', 'tissue', 'hdmi', 'gc for interim', 'pens', 'file storage', 'sharpies', 'broom', 'paper towels', 'tape', 'trash bags', 'copies', 'key tags', 'shelf clips', 'network switch', 'brochure holder', 'file folders', 'organizer', 'box for', 'laminate paper', 'folder', 'hanger', 'hooks', 'milk crates', 'grill cover', 'devices to help speak', 'umbrella', 'folders', 'trash liners', 'sunscreen', 'air fresheners', 'minor information', 'mailbox', 'prints', 'organization supplies', 'radios', 'radio', 'shoe boxes', 'storage box', 'clip boards', 'color print of loading keypad', 'toilet plunger', 'dewalt', 'gc for site', 'ethernet cable', 'key ring', 'label maker', 'bath tissue', 'AAs', 'wall plug', 'office chargers', 'folding chair', 'lights', 'air freshener', 'legal pads', 'container', 'reimbursement for team members purchasing supplies','new flag and reroped the pole','brackets hanging','organizers', 'table', 'drawer organizers','binder', 'page protector', 'lamination pouches','pants','usb travel kit','hangers','xpt items','paper','Mirrors fod hitches and beds of trucks','chargers', 'charging block','batteries for site','Walkie Talkies','Plunger', 'note pads', 'calendar, mousepad','Charging block', 'swiffer refills,','Computer charger','speakers','Binders and 3 compartments drawer','Thermal paper','organizers','envelopes','walkie talkies','Dog treat containers','collapsable stool for XPT', 'member has dr note to sit as needed','paper for xpt','2 monitors for loading area','envelopes','flags for memorial Day and 4th of july', 'scent plug ins', 'hand soap', 'rain ponchos', 'locking file', 'post its', 'keyboards', 'leadership book', 'visa for asls', 't.p.', 'restroom art', 'mail boxes', 'filing', 'Storage supplies', 'brass hooks', 'rain suits', 'Garbage Can Inserts', 'Organizational equipment', 'Manager communication book', 'headphones', 'laminating supplies', 'sun screen', 'Ports to fix cables in booth area', 'Kleenex', 'Lock Box', 'walkies', 'Key board', 'it cables', 'internet cable'],
    ###################################################################################
    'Parking': ['parking'],
    ################################## Parts ##########################################
    'Parts': ['fittings', 'fitting','bolts', 'parts', 'valve', 'screws', 'shaft sleeve', 'couplers', 'fasteners', 'wire nuts', 'grating', 'hydraulic fittings', 'quick connects', 'steel weld', 'v belts', 'key stock', 'bottom cap', 'hydraulic hose', 'elbow', 'hose for site', 'wheel for', 'hydraulic hoses', 'vfd for', 'valves', 'quick connects', 'door seal', 'doors stop', 'quick connect', 'nipple connector', 'hose barb', 'hose clamps', 'gfi', 'extra wheels', 'parts for', 'tapcons for', 'union joints', 'wire to test', 'to repair', 'washers', 'door stops', 'nipples', 'restricters', 'poly tubing', 'ball valve', 'tubing', 'anchors', 'adapters', 'foaminator air manifold','adapter', 'barb, pipe nipple','UV lamp/pressure switch','High pressure hose replacement','Mac Val for buffing shin','Rotator bearing assembly','towel bar to hang bottles','hydroflex hose','Compressor fuses','tapcon','Membrain','Membrain','Door knob', 'stand','Shocks, strut', 'springs caps','Repairs for broken Lines','Hex nuts','bolt','GFCI replacement','GFCI replacement','GFCI replacement','clevis and pin for top brush 1','Hose repair','brass barb', 'pressure washer repair', 'belts', 'part for', 'sonar', 'hose seals', 'Electrical outlet', 'replacement belt', 'nuts to fix', 'heater hose', 'fuse', 'Gaskets', 'Couplings', 'Connectors', 'vpd drive', 'fastners', 'tank lids', 'flange', 'nipple', 'flexe hose', 'pw gasket', 'hp hose', 'refund for spare hose','Quick coupler','O rings','door hinges','door handles', 'grates for tunnel'],
    ###################################################################################
    'Permits': [],
    'Pest Control/Extermination': ['fire ant killer', 'insect killer', 'pest control', 'ant bait', 'gopher', 'bird repellents', 'ant killer', 'rat bait', 'bug', 'repellant', 'pest','arrow', 'exterminators', 'wasp spray'],
    'Pit Pumping': ['pit', 'waders', 'pumped out', 'wader'],
    ###################### Plumbing Expense ###########################################
    'Plumbing Expense': ['plumbing', 'pvc', 'reclaim', 'piping', 'pipe to move', 'sharkbites', 'plumber', 'sharkbite', 'pipe replacement'],
    ##################################################################################
    'Postage': ['shipping', 'postage', 'mail spinners', 'shrink wrap', 'stamp'],
    'Professional Fees': ['Fine from City'],
    'Professional Fees - Non-SL': [],
    'Professional Memberships': ['chamber'],
    'Recruitment': ['indeed', 'ad for asl'],
    'Relocation': [],
    'Rental Equipment': ['warehouse space', 'storage unit', 'truck rental', 'powerwasher rental', 'storage building', 'storage'],
    'Repairs': [],
    'Rewrite Memo': ['lowes', 'wash supplies'],
    'Safety Supplies': ['gloves', 'emergency', 'fire extinguisher', 'hazardous signage', 'railing', 'safety signage', 'cooling wraps', 'knee pads', 'masks', 'guard rail', 'diesel cans', 'junction box covers', 'fire resistant ceiling tiles', 'cones', 'fire ext.', 'safety supplies'],
    'Security': ['padlocks', 'pad locks', 'locks', 'door lock'],
    'Signage': ['signs', 'velcro for signage', 'lettering', 'sign', 'sandbags', 'posts for signage', 'lock for dumpster'],
    ################################ Small Tools ######################################
    'Small Tools': ['fan', 'jump starter', 'blades', 'wire stripper', 'wench puller', 'small tools', 'drill', 'drain snake', 'pipe cutter', 'tools', 'scrapper', 'funnel', 'fans', 'small tools', 'come along', 'ratchet straps', 'wrench', 'clamps', 'transfer pump', 'shop vacuum', 'ladder', 'bungees', 'dolly', 'water key', 'floor jack', 'socket set', 'crowbar', 'file tool', 'grinding wheel', 'sawzall blade', 'leaf blower', 'pressure washer hose', 'pressure washing hose', 'bearing puller', 'pressure washing wand', 'bits', 'tie down straps', 'sockets', 'hex wrenches', 'screwdriver', 'tool hangers', 'wrenches', 'pressure wash hoses', 'hose reel', 'pressure guns','pressure washer hoses', 'pressure washer lines', 'sprayer', 'spray nozzles', 'glass scrapper', 'foam cannon', 'grinder', 'auger', 'hook and pick set', 'caulk gun', 'reciprocating saw', 'pliers', 'ratchets', 'screw extractor', 'pressure washer gun', 'blade knives', 'battery charger','Pressure washing tips','tool holder', 'driver set','hook magnet','Shop Vac','Power washer hose','Hydraulic crimp','Powerwasher hoses','Ryobi charger','Ryobi batterys', 'Toolbox for xpt','rumble strips and cutting blade','Purchased a new ruler for the chemical barrel  measurement','Orbital scrubber for windows','Tool box','Returned rivet gun','torch', 'pressure washer nozzles', 'foam gun', 'hook and pick set', 'putty knives', 'drum pump', 'ratchet strap', 'scrappers', 'foam cannons', 'hand pump', 'sewer grate opener', 'ratchet set', 'barrel pump', 'step stool', 'pitcher for fluid filling', 'measuring stick', 'Brush sticks', 'Brushes', 'cable puller', 'tip cleaners', 'Hex Keys', 'Sawzall', 'rake', 'screw drivers', 'come-along','comealong','Strap puller','Chemical ruler', 'Grinding blade', 'extension cord', 'extension cords'],
    ################################ Snacks ###########################################
    'Snacks': ['water bottles', 'bottle water','gatorade', 'drinks', 'powerade', 'water for', 'bottled water', 'water btl', 'waters for', 'snacks for', 'btl water','gatorades for', 'bottles of water', 'ice', 'gatorades', 'water flavoring', 'popsicles', 'electrolyte', 'water jug', 'primo water', 'muffins', 'snacks', 'snack', 'propel', 'water jugs', 'water refill', 'ice for employees', 'drink mix', 'powerades', 'water cases', 'waters', 'case of water'],
    ###################################################################################
    'Snow removal': ['salt application'],
    'Supplies - Non-SL': [],
    'Taxi and Ground Travel': ['uber', 'car service from'],
    'Telephone & Internet': [],
    'Tesla Charger': [],
    'Tolls': ['tolls', 'toll'],
    'Training': ['for training', 'handbooks'],
    'Travel Meals': ['per diem', 'travel meal', 'carwash college meal', 'travel dinner', 'groceries for travel'],
    'Uniforms': [],
    ################################### Vehicle Damage Claims ##############################
    'Vehicle Damage Claims': ['car towed', 'sunroof guard', 'luggage racks', 'side mirror', 'wheel covers', 'replacement for bug shield', 'wiper blade', 'wiper replacement', 'tow truck', 'wipers', 'center cap', 'center caps', 'damage claims', 'antenna', 'members car', 'Mirror for Porsche', 'Camera for Mitsubishi', 'car damage', 'broken wiper', 'emblem'],
    #################################### Wash Maintenance ##################################
    'Wash Maintenance': ['oil', 'hose hangers', 'cleaning supplies', 'detergent', 'degreaser', 'concrete', 'filter', 'dawn', 'vinegar', 'salt', 'strainer', 'paint', 'hydraulic oil', 'tide pods', 'roof sealer', 'grease', 'scrub pads', 'zip ties', 'tunnel cleaner', 'cleaners for', 'hydraulic fluid', 'storage container', 'water hoses', 'Towel coolers', 'wd40', 'pink stuff', 'mop', 'storage containers', 'degreasers', 'laundry', 'rust remover', 'scrub brush', 'swiffer pads', 'squeegee', 'spill absorb', 'cleaning chemicals', 'aluminum cleaner', 'antifreeze', '2x4\'s', 'air filters', 'chemical spill','extra dumpster', 'glass cleaner', 'water hose', 'filters', 'filters for', 'zep', 'flex paste for tunnel','driveway cleaner', 'hand cleaner', 'cleaning solution', 'aw32', 'flex seal', 'pb blaster', 'sheet metal', 'wall cleaner', 'purple power', 'alumabrite', 'silicon', 'all purpose cleaner', 'wool pads', 'Shovel and bar cleaner wall','utility mat', 'gain','Lube for chemical equipment', 'Hitch Mirror', 'Prep Hose','Power washer reel','Hose reels','Cleaners/chemicals for tunnel cleaning','storage totes for towels','furniture polish','Product to clean a seal windows in the tunnel','Supplies to clean aluminum arches in the tunnel','C3 for cleaning arches','Trash can and towel cart','Towel carts and trash cans for towels','2 trash cans','pot hole repair','glue', 'anti freeze', 'c3', 'water softener', 'propane', 'cleaning products', 'dish soap', 'ajax', 'steel wool', 'bleach', 'lime off', 'to landfill', 'storage tote', 'water softner salts', 'shop rags', 'WD-40', 'window mops', 'goo gone', 'CLR for car wash', 'germx', 'purex', 'alumanew', 'magic erasers', 'soap for windows', 'water softner pellets', 'anti sieze', 'clr', 'cleaner', 'cleaning chems', 'oxiclean', ],
    #########################################################################################
    'Wash Repairs': ['fee to diagnose', 'pump fixed', 'compressor installation', 'pump repair','repair on motor', 'dd/br services', 'exit plate welded', 'Repairs labor', 'rolling door 3rd party repair', 'pump repair', 'vacuum motor install', 'weld gate', 'fix entrance door'],
    'Wash Repairs - Non-SL': [],
    #################################### Wash Supplies ######################################
    'Wash Supplies': ['microfiber', 'bug sponges', 'spray bottles', 'bug blocks', 'bug scrubbing pads', 'prep buckets',  'bucket', 'grill gates', 'grill grates', 'brush heads', 'bug scubbing pads', 'brush handles', 'wax', 'buckets', 'prep carts', 'cart for xpt', 'sponges', 'xpt cart', 'bug scrubbers', 'prep brushes', 'deck brushes', 'prep soap', 'bug remover', 'towels', 'spray bottles', 'triggers'],
    #########################################################################################
    'Water & Sewer Repairs': ['repair to water main', 'water main break'],
}

# Define a function to categorize the expenses based on the 'Notes' column using the dictionary
def categorize_expense(note, category_dict):
    note = str(note)
    note_lower = note.lower()
    matched_categories = []
    for category, keywords in category_dict.items():
        for keyword in keywords:
            # Use regular expression to match whole words only
            if re.search(r'\b' + re.escape(keyword.lower()) + r'\b', note_lower):
                matched_categories.append(category)
                break # Once a match is found, no need to check further keywords for this category
    if len(matched_categories) == 1:
        return matched_categories[0], matched_categories
    elif len(matched_categories) > 1:
        return 'Itemized', matched_categories
    else:
        return 'Manual', matched_categories

# Apply the function to create new columns in the DataFrame
categorized_data = df['Notes'].apply(lambda note: categorize_expense(note, category_dict))
if len(categorized_data) == len(df):
    df['New'] = pd.Series(categorized_data.tolist(), index=df.index)
else:
    print("Mismatch in lengths: df and categorized_data")
# Load the workbook and the specific sheet to update
wb = load_workbook(data_path)
ws = wb['Need New']
# ws = wb['Cats']

# Update the 'New' and 'Matched Cats' columns in the Excel sheet
for index, row in df.iterrows():
    ws.cell(row=index + 2, column=6, value=row['New'][0])
    # ws.cell(row=index + 2, column=3, value=row['New'])
    # Assuming 'Matched Cats' is in column K (11th column, index 10 in 0-based index)
    # matched_cats = row['Matched Cats'] if isinstance(row['Matched Cats'], list) else []
    # ws.cell(row=index + 2, column=10, value=', '.join(matched_cats))

# Save the workbook
wb.save(data_path)

print("Updated Excel file with new 'New' columns without removing data validation or format.")
