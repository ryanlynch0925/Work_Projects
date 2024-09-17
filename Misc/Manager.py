import openpyxl
import win32com.client as win32
from datetime import datetime
import time

# Record the start time
start_time = time.time()

# Read data from Excel workbook
workbook = openpyxl.load_workbook(r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\Master Expense Report V4.0.xlsx", data_only=True)
s_sheet = workbook['Summary']
aa_sheet = workbook['Awaiting Approval Data']

# print(f"Active sheet name: {s_sheet.title}")

# Define the last row in column B
s_iLastRow = s_sheet.max_row
aa_iLastRow = aa_sheet.max_row

# Create the Outlook application object
outlook = win32.Dispatch('Outlook.Application')

manager_data = {}

# Loop through the rows in column B
for i in range(5, s_iLastRow + 1):
    calculated_value = s_sheet.cell(row=i, column=22).value
    if calculated_value is None or calculated_value == '':
        continue
    elif calculated_value > 0:
        manager = s_sheet.cell(row=i, column=32).value
        employee = s_sheet.cell(row=i, column=1).value
        amount = round(float(s_sheet.cell(row=i, column=22).value), 2)
        
        # Create or update the manager's data
        if manager not in manager_data:
            manager_data[manager] = {'total_amount': 0, 'expenses': []}
        manager_data[manager]['total_amount'] += amount
        manager_data[manager]['expenses'].append({
            'employee': employee,
            'amount': amount
        })

# Loop through the "Awaiting Approval Data" sheet
for aa_row in range(2, aa_iLastRow + 1):
    aa_employee = aa_sheet.cell(row=aa_row, column=6).value
    aa_manager = aa_sheet.cell(row=aa_row, column=10).value
    expense_report = aa_sheet.cell(row=aa_row, column=1).value
    expense_amount = aa_sheet.cell(row=aa_row, column=8).value
    
    for manager, data in manager_data.items():
        for expense in data['expenses']:
            if expense['employee'] == aa_employee and aa_manager == manager:
                expense['reports'] = expense.get('reports', [])
                expense['reports'].append({
                    'report': expense_report,
                    'amount': expense_amount
                })
            

# Loop through the aggregated data and send emails
for manager, data in manager_data.items():
    mail = outlook.CreateItem(0)  # 0 represents an email item
    mail.To = manager
    mail.CC = "Leigh Stalling; Kevin McGonigle; Travis Powell"
    mail.Subject = "Expenses awaiting Manager Approval"
    
    email_body = f"Dear {manager},<br><br>"
    email_body += 'Please navigate to your <span style="color: #72C596"><b>WorkDay Inbox</b></span> to review and approve report(s) within 7 days.<br><br>'
    email_body += "The following employee(s) are awaiting your approval of their expenses:<br><br>"
    
    
    for expense in data['expenses']:
        email_body += f"<b>&nbsp;&nbsp;{expense['employee']}: <span style='color: #003896'>${expense['amount']:.2f}:<br></b></span>"
        if 'reports' in expense:
            for report in expense['reports']:
                email_body += f"&nbsp;&nbsp;&nbsp;&nbsp;--Expense Report: {report['report']}<br>"
    
    signature = '''
    <br><span style= 'color: #E476F44; font-size: 22pt'>David Ryan Lynch</b><br></span>
    PH: +1 706-481-2635<br>
    T & E Specialist<br>
    Home Office
    '''
    email_body += signature
    mail.HTMLBody = f"<html><body>{email_body}</body></html>"
    mail.Display()
    mail.Send()
# Record the end time
end_time = time.time()

    # Calculate the elapsed time
elapsed_time = end_time - start_time

print(f"Total execution time: {elapsed_time:.2f} seconds")