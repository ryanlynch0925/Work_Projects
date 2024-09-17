import openpyxl
import win32com.client as win32
from datetime import datetime
import time

# Record the start time
start_time = time.time()

# Read data from Excel workbook
workbook = openpyxl.load_workbook(r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Lisa Project.xlsx", data_only=True)
#workbook = openpyxl.load_workbook(r"C:\Users\LisaHughes\OneDrive - Tidal Wave Autospa\Desktop\Car Wash Services\Lisa Project.xlsx", data_only=True)
sheet = workbook['Site Info']

#print(f"Active sheet name: {sheet.title}")

# Define the last row in column B
iLastRow = sheet.max_row
#print(iLastRow)

# Create the Outlook application object
outlook = win32.Dispatch('Outlook.Application')

batch_emails = []
month = 'August'
due_date = '8/31/2023'
due_time = '3:00 PM'
month_folder = '2023-07'

for i in range(2, iLastRow +1):
    site = sheet.cell(row=i, column=1).value
    print(site)
    site_name = sheet.cell(row=i, column=2).value
    site_leader = sheet.cell(row=i, column=3).value
    site_email = sheet.cell(row=i, column=4).value
    consultant_name = sheet.cell(row=i, column=5).value
    if site is not None:
        mail = outlook.CreateItem(0)  # 0 represents an email item

        # Set the email properties
        mail.To = f'{site_email}'
        mail.CC = f'{consultant_name}'
        mail.Subject = f"{month} Invoices for {site_name}"
        attachment_path1 = rf"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\{site}.approval.xlsx"
        print(attachment_path1)
        attachment_path2 = rf"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\{site}.pdf"
        print(attachment_path2)
        # attachment_path1 = rf"C:\Users\LisaHughes\OneDrive - Tidal Wave Autospa\Desktop\Car Wash Services\{month_folder}\{site}.xlsx"
        # attachment_path2 = rf"C:\Users\LisaHughes\OneDrive - Tidal Wave Autospa\Desktop\Car Wash Services\{month_folder}\{site}.pdf"
        mail.Attachments.Add(attachment_path1)
        mail.Attachments.Add(attachment_path2)

        signature = '''
        <b style="font-size: 22pt; color: black;">Lisa Hughes</b><br>
        <b style="font-size: 11pt; color: #333;">Accounts Payable Specialist</b><br>
        <span style="font-family: 'Centaur', Georgia, serif; font-size: 14pt; color: #548DD4;">Tidal Wave Management LLC</span><br>
        <span style="font-size: 12pt; color: black;">Post Office Box 311</span><br>
        <span style="font-size: 12pt; color: black;">115 East Main Street</span><br>
        <span style="font-size: 12pt; color: black;">Thomaston GA 30286</span><br>
        <span style="font-size: 12pt; color: black;">(o) 706-647-0414 ext. 147</span><br>
        <span style="font-size: 12pt; color: black;">(d) 706-646-7756</span><br>
        <span style="font-size: 12pt; color: black;">(f) 706-647-0474</span><br>

        '''
        emailBody = f'''Dear {site_leader}<br><br>

        Please review these recent chemical invoices attached, fill in spreadsheet and return by {due_time} {due_date}.<br><br>

        '''
        emailBody += signature
        mail.HTMLBody = f"<html><body>{emailBody}</body></html>"
        batch_emails.append(mail)
    else:
        pass

for email in batch_emails:
    email.Display()
    #email.Send()
# Record the end time
end_time = time.time()

    # Calculate the elapsed time
elapsed_time = end_time - start_time

print(f"Total execution time: {elapsed_time:.2f} seconds")