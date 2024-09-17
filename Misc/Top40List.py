import openpyxl
import win32com.client as win32
from datetime import datetime
import time
import cProfile

# Record the start time
start_time = time.time()


# Read data from Excel workbook
workbook = openpyxl.load_workbook(r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\Master Expense Report V4.0.xlsx", data_only=True)
sheet = workbook['Summary']

iLastRow = sheet.max_row

outlook = win32.Dispatch('Outlook.Application')

top_40_records = []

for i in range(5, iLastRow + 1):
    # Check if the rank is less than or equal to 40
    calculated_value = sheet.cell(row=i, column=27).value
    if calculated_value is None or calculated_value == '':
        pass
        #print(f"Value at row {i}, column 27 is None or blank")
    elif isinstance(calculated_value, (int, float)):
        rank = int(calculated_value)
        if rank <= 40:  # Column AA = 27
            #print(f"Processing row {i}, rank: {rank}")
            outstanding = round(float(sheet.cell(row=i, column=25).value))
            employee = sheet.cell(i, 1).value
            Top40Outstanding = round(float(sheet.cell(2, 26).value))
            Top40Percentage = float(sheet.cell(1,26).value)
            top_40_records.append((rank, employee, outstanding, Top40Outstanding, Top40Percentage))

# Create the email
mail = outlook.CreateItem(0)  # 0 represents an email item
    # Set the email properties
mail.To = 'Leigh Stallings; Kevin McGonigle'
mail.CC = f"Keri Pack; Karla Kendrick"
mail.Subject = "Top 40 Outstanding List"
#mail.Subject = "Testing Automatic Oustanding Emails"
emailBody = 'Top 40 Outstanding List, to look at on Monday:<br>'
emailBody += f'<br>Top 40 Outstanding Balance: <b>${Top40Outstanding:,.2f}</b><br>'
emailBody += f'Top 40 Outstanding Percentage: <b>{Top40Percentage*100:.2f}%</B><br><br>'
for rank, employee, outstanding, Top40Outstanding, Top40Percentage in top_40_records:
    emailBody += f"{rank}: {employee}, <b>${outstanding:,.2f}</b><br>"

mail.HTMLBody = f"<html><body>{emailBody}</body></html>"

#Display the Email
mail.Display()

# Send the email
#mail.Send()

# Record the end time and calculate the execution time
end_time = time.time()
execution_time = end_time - start_time
print(f"Execution time: {execution_time} seconds")