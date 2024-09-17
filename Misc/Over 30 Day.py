import openpyxl
from openpyxl import load_workbook
import win32com.client as win32
import pandas as pd
import traceback
from datetime import datetime

try:
    outlook = win32.Dispatch('Outlook.Application')
except Exception as e:
    print(f"Error occurred while connecting to Outlook: {e}")
    traceback.print_exc()  # Print detailed traceback for debugging

excel_file = r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\Master Expense Report V4.0.xlsx"
df = pd.read_excel(excel_file, sheet_name='Data')

# Filter rows where 'Days Old' is greater than or equal to 30
filtered_df = df[(df['Days Old'] >= 30)]
# filtered_df = df[(df['Days Old'] >= 30) & ((df['Expense Report Status Detail'] == '') | (df['Expense Report Status Detail'] == 'Sent Back'))]

# Define columns to remove
columns_to_remove = ['Credit Card', 'Credit Card Transaction', 'Charge Date', 'Expense Report Date',
                     'Report Days Old', 'Prior?']

# Drop unnecessary columns
filtered_df = filtered_df.drop(columns=columns_to_remove)
filtered_df = filtered_df.fillna('')

# Initialize a dictionary to store combined expenses for each employee
combined_expenses = {}

# Iterate through the filtered DataFrame
for index, row in filtered_df.iterrows():
    employee = row['Employee']
    amount = row['Amount']

    # Skip certain locations
    if row['Location'] in ['SHJ Construction LLC', 'Stangood-GA', 'Stangood-OH', 'Terminated']:
        continue

    # If employee not in dictionary, add them
    if employee not in combined_expenses:
        combined_expenses[employee] = {
            'total_amount': 0,
            'num_transactions': 0,
            'transactions': []
        }

    # Add transaction details to the dictionary
    combined_expenses[employee]['total_amount'] += amount
    combined_expenses[employee]['num_transactions'] += 1
    # combined_expenses[employee]['transactions'].append({
    #     'amount': amount,
    #     'load_date': row['Load Date'],
    #     'expense_report': row['Expense Report'],
    #     'expense_report_status': row['Expense Report Status'],
    #     'expense_report_status_detail': row['Expense Report Status Detail'],
    # })

# Specify the Excel file to update
output_excel_file = r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\30-Day Report\30-Day Report.xlsx"

# Check if the output file already exists
try:
    wb = load_workbook(output_excel_file)
except FileNotFoundError:
    # Create a new workbook if the file doesn't exist
    wb = openpyxl.Workbook()

# Get today's date as a string
today_date = datetime.now().strftime('%m-%d-%Y')

# Sort the combined_expenses dictionary by 'total_amount' in descending order
sorted_expenses = dict(sorted(combined_expenses.items(), key=lambda x: x[1]['total_amount'], reverse=True))

# Create a new sheet with today's date as the title
ws = wb.create_sheet(title=today_date)

# Write headers if the sheet is empty
if ws.max_row == 1 and ws.max_column == 1:
    ws.append(['Employee', 'Total', 'Number of Transactions'])

# Iterate through the combined_expenses dictionary
for employee, details in sorted_expenses.items():
    # Append data to the sheet
    ws.append([employee, details['total_amount'], details['num_transactions']])

# Save the workbook
wb.save(output_excel_file)
print(f"Data has been added to {output_excel_file} on the sheet '{today_date}'")