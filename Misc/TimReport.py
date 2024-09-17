import openpyxl
import win32com.client as win32
from datetime import datetime
import time
import cProfile

# Record the start time
start_time = time.time()


# Read data from Excel workbook
workbook = openpyxl.load_workbook(r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\Master Expense Report V4.0.xlsx", data_only=True)
sheet = workbook['Summary']

iLastRow = sheet.max_row

outlook = win32.Dispatch('Outlook.Application')

top_40_records = []

for i in range(5, iLastRow + 1):
    # Check if the rank is less than or equal to 40
    Tim = sheet.cell(row=i, column=34).value
    Percentage = sheet.cell(1, 34).value
    TimTotal = sheet.cell(2, 34).value
    if Tim == 'Yes':
        #print(f"Processing row {i}, rank: {rank}")
        rank = sheet.cell(row=i, column=27).value
        outstanding = round(float(sheet.cell(row=i, column=25).value))
        employee = sheet.cell(i, 1).value
        UT = float(sheet.cell(i, 18).value)
        ID = float(sheet.cell(i, 19).value)
        IP = float(sheet.cell(i, 20).value)
        if outstanding > 0:
            top_40_records.append((rank, employee, outstanding, UT, ID, IP))

# Create the email
mail = outlook.CreateItem(0)  # 0 represents an email item
    # Set the email properties
mail.To = 'Timothy Fruge'
mail.CC = f"Bruce Maxwell; Leigh Stallings"
mail.Subject = "Operations Outstanding Expense List"
#mail.Subject = "Testing Automatic Oustanding Emails"
emailBody = 'Operations Outstanding Expense List:<br><br>'
emailBody += f'Percentage of Outstanding Total: <b>{Percentage*100:.2f}%</b><br>'
emailBody += f'Total of Outstanding Total: <b>${TimTotal:,.2f}</b><br>'

for rank, employee, outstanding, UT, ID, IP in top_40_records:
    emailBody += f"{rank}: {employee}, <b>${outstanding:,.2f}</b><br>"
    emailBody += f'''
                    <ul>
                        <li>Not Submitted: <b>${UT:,.2f}</b></li>
                        <li>In Draft: <b>${ID:,.2f}</b></li>
                        <li>In Progress: <b>${IP:,.2f}</b></li>
                    </ul>
                    '''

mail.HTMLBody = f"<html><body>{emailBody}</body></html>"

#Display the Email
mail.Display()

# Send the email
#mail.Send()

# Record the end time and calculate the execution time
end_time = time.time()
execution_time = end_time - start_time
print(f"Execution time: {execution_time} seconds")