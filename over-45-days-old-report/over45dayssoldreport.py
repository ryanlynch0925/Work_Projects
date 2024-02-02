import pandas as pd
from openpyxl import Workbook, load_workbook
import os

class ExpenseReportProcessor:
    """
    This class is responsible for processing expense reports. It reads in the data from an Excel file, filters and cleans the data, generates a summary of the data, and saves the summary to an Excel file.

    Parameters
    ----------
    data_path : str
        The path to the Excel file containing the data.
    summary_file_path : str
        The path to the Excel file where the summary should be saved.

    Attributes
    ----------
    data_path : str
        The path to the Excel file containing the data.
    summary_file_path : str
        The path to the Excel file where the summary should be saved.
    columns_to_remove : list
        A list of columns that should be removed from the data.

    """

    def __init__(self, data_path, summary_file_path):
        self.data_path = data_path
        self.summary_file_path = summary_file_path
        self.columns_to_remove = ['Credit Card', 'Credit Card Transaction', 'Charge Date', 'Expense Report Date',
                                    'Report Days Old', 'Prior?']

    def read_data(self, sheet_name='Data'):
        """
        Reads in the data from the Excel file.

        Parameters
        ----------
        sheet_name : str, optional
            The name of the sheet containing the data. The default is 'Data'.

        Returns
        -------
        pandas.DataFrame
            The data from the Excel file.

        """
        if not os.path.exists(self.data_path):
            raise ValueError('Data file not found')
        return pd.read_excel(self.data_path, sheet_name=sheet_name)

    def filter_and_clean_data(self, df):
        """
        Filters and cleans the data.

        Parameters
        ----------
        df : pandas.DataFrame
            The data to be filtered and cleaned.

        Returns
        -------
        pandas.DataFrame
            The filtered and cleaned data.

        """
        filtered_df = df[df['Days Old'] >= 45].copy()
        filtered_df = filtered_df.drop(columns=self.columns_to_remove)
        filtered_df = filtered_df.fillna("Not Submitted")
        return filtered_df

    def generate_summary(self, final_df):
        """
        Generates a summary of the data.

        Parameters
        ----------
        final_df : pandas.DataFrame
            The filtered and cleaned data.

        Returns
        -------
        pandas.DataFrame
            The summary of the data.

        """
        unique_employees = final_df['Employee'].unique()
        all_summaries = []

        for employee in unique_employees:
            employee_df = final_df[final_df['Employee'] == employee]
            manager = employee_df['Manager'].iloc[0]  # Get the manager for the current employee
            over_45_total = employee_df['Amount'].sum()
            not_submitted_total = employee_df.loc[employee_df['Expense Report Status Detail'] == 'Not Submitted', 'Amount'].sum()
            sent_back_total = employee_df.loc[employee_df['Expense Report Status Detail'] == 'Sent Back', 'Amount'].sum()

            # Create a summary DataFrame for the current employee
            employee_summary = pd.DataFrame({
                'Employee': [employee],
                'Over 45 Days (Total)': [over_45_total],
                'Sent Back (Total)': [sent_back_total],
                'Not Submitted (Total)': [not_submitted_total],
                'Waiting on Manager (Total)': [0],  # Assuming there's no waiting in your current logic
                'Manager': [manager]  # Replace with the actual manager information if available
            })

            all_summaries.append(employee_summary)

        final_summary_df = pd.concat(all_summaries, ignore_index=True)
        final_summary_df.loc[:, 'Waiting on Manager (Total)'] = 0
        return final_summary_df

    def save_summary_to_excel(self, final_summary_df):
        """
        Saves the summary to the Excel file.

        Parameters
        ----------
        final_summary_df : pandas.DataFrame
            The summary of the data.

        """
        if not os.path.exists(self.summary_file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Summary Report'

            with pd.ExcelWriter(self.summary_file_path, engine='openpyxl') as writer:
                final_summary_df.to_excel(writer, sheet_name='Summary Report', index=False)
        else:
            book = load_workbook(self.summary_file_path)

            with pd.ExcelWriter(self.summary_file_path, engine='openpyxl', if_sheet_exists='replace', mode='a') as writer:
                final_summary_df.to_excel(writer, sheet_name='Summary Report', index=False, header=True)

# Example usage
data_path = r'c:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\Test.xlsx'
summary_file_path = r"C:\Users\DavidLynch\OneDrive - Tidal Wave Autospa\Documents\30-Day Report\45-Day Report.xlsx"

report_processor = ExpenseReportProcessor(data_path, summary_file_path)
data_df = report_processor.read_data()
filtered_data_df = report_processor.filter_and_clean_data(data_df)
final_summary_df = report_processor.generate_summary(filtered_data_df)
report_processor.save_summary_to_excel(final_summary_df)